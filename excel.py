from openpyxl import load_workbook
from openpyxl.styles import Alignment

def rellenar_cargos_excel(ruta_archivo, cargos):
    """
    Rellena un archivo Excel (.xlsx) con una lista de cargos de manera dinámica y vertical en celdas combinadas,
    ajustando el ancho de las columnas.
    
    Parámetros:
    - ruta_archivo: Ruta del archivo Excel .xlsx a modificar.
    - cargos: Lista de cargos que se van a agregar.
    """
    try:
        # Cargar el archivo Excel
        wb = load_workbook(ruta_archivo)
        sheet = wb.active  # Seleccionar la hoja activa

        # Configurar la posición inicial (G10 es columna 7, fila 10)
        columna_inicial = 7  # Columna G
        fila_inicial = 10

        for idx, cargo in enumerate(cargos):
            columna_actual = columna_inicial + idx
            letra_columna = sheet.cell(row=1, column=columna_actual).column_letter  # Obtener letra de la columna

            # Ajustar el ancho de la columna
            sheet.column_dimensions[letra_columna].width = 3  # Ancho de 30 píxeles

            # Identificar la celda principal si está combinada
            celda_titulo = sheet.cell(row=fila_inicial, column=columna_actual)

            for merged_range in sheet.merged_cells.ranges:
                if celda_titulo.coordinate in merged_range:
                    # Si la celda está en un rango combinado, obtener la celda principal
                    celda_titulo = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break

            # Escribir el cargo y configurarlo en orientación vertical
            celda_titulo.value = cargo
            celda_titulo.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")

        # Guardar los cambios en el archivo
        wb.save(ruta_archivo)
        print("Archivo actualizado correctamente.")

    except Exception as e:
        print(f"Error al actualizar el archivo: {e}")

# Ruta del archivo y lista de cargos
ruta_archivo = "C://Users//jozek//Documents//5.xlsx"
cargos = ["Cargo 1", "Cargo 2", "Cargo 3", "Cargo 4"]

rellenar_cargos_excel(ruta_archivo, cargos)
