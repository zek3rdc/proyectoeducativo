import os
from datetime import datetime, timedelta
import calendar
import openpyxl
from openpyxl.utils import get_column_letter
import psycopg2
from psycopg2 import OperationalError

class ReporteExcel:
    def __init__(self, plantilla_path, output_path):
        self.plantilla_path = plantilla_path
        self.output_path = output_path

    @staticmethod
    def valor_o_cero(valor):
        """Devuelve 0 si el valor es None, vacío o null, de lo contrario devuelve el valor."""
        return valor if valor not in (None, '', 'null') else 0

    def rellenar_plantilla_excel(self, data, fecha_inicio=None, fecha_fin=None):
        # Eliminar el archivo de salida si ya existe
        if os.path.exists(self.output_path):
            os.remove(self.output_path)

        # Calcular el rango de fechas si no se proporcionan
        if not fecha_inicio or not fecha_fin:
            today = datetime.today()
            fecha_inicio = today.replace(day=1)  # Primer día del mes actual
            fecha_fin = today.replace(day=28) + timedelta(days=4)  # Último día del mes actual
            fecha_fin = fecha_fin - timedelta(days=fecha_fin.day)  # Ajustar al último día del mes

        # Calcular el número de días del rango
        dias_rango = (fecha_fin - fecha_inicio).days + 1  # Incluye ambos días en el cálculo

        # Abrir el archivo de la plantilla
        wb = openpyxl.load_workbook(self.plantilla_path)
        sheet = wb["PZ Nombre de la Institución "]  # Asegurar el nombre correcto de la hoja

        # Colocar el número de días en la fila AI5
        sheet["AI5"] = dias_rango

        # Obtener el nombre del mes y colocarlo en AC4
        nombre_mes = calendar.month_name[fecha_inicio.month]
        meses_esp = {
            "January": "Enero", "February": "Febrero", "March": "Marzo", "April": "Abril", "May": "Mayo",
            "June": "Junio", "July": "Julio", "August": "Agosto", "September": "Septiembre", "October": "Octubre",
            "November": "Noviembre", "December": "Diciembre"
        }
        sheet["AC4"] = meses_esp[nombre_mes]

        fila = 12  # Fila inicial para los datos

        # Detectar dinámicamente los cargos únicos
        cargos_unicos = sorted(set(cargo for prof in data if prof.get("cargo") for cargo in prof["cargo"]))

        # Asignar dinámicamente columnas a los cargos
        columna_inicio = 7  # Columna "G" en Excel (índice base 1)
        cargos_columnas = {cargo: get_column_letter(columna_inicio + i) for i, cargo in enumerate(cargos_unicos)}

        # Escribir los nombres de los cargos en la fila 10
        for cargo, col in cargos_columnas.items():
            sheet[f"{col}10"] = cargo

        # Mapeo de columnas para Codificaciones y Categorías
        codificaciones = {"Lic": "Z", "PG": "AA", "PGE": "AB", "TSU": "AC", "Br.Dc.": "AD", "NG": "AE"}
        categorias = {"I": "AF", "II": "AG", "III": "AH", "IV": "AI", "V": "AJ", "VI": "AK"}

        for index, prof in enumerate(data, start=1):
            row = fila + index - 1  # Fila en la que se insertará la información

            # Índice en la columna A
            sheet[f"A{row}"] = index

            # Nombre completo en la columna B
            sheet[f"B{row}"] = f"{prof['nombre']} {prof['apellido']}"

            # Cédula en la columna C
            sheet[f"C{row}"] = prof['cedula']

            # Fecha laboral en la columna E
            fecha_laboral = prof.get('fecha_laboral')
            if fecha_laboral:
                if isinstance(fecha_laboral, datetime):
                    fecha_laboral = fecha_laboral.date()  # Convertir a datetime.date si es datetime
                diferencia_anios = (datetime.now().date() - fecha_laboral).days // 365
            else:
                diferencia_anios = ""  # Dejar la celda vacía

            sheet[f"E{row}"] = diferencia_anios

            # Marcar los cargos en sus respectivas columnas
            if prof.get("cargo"):
                for cargo in prof["cargo"]:
                    if cargo in cargos_columnas:
                        sheet[f"{cargos_columnas[cargo]}{row}"] = "X"

            # Estudios (Sí/No en T y U)
            sheet[f"T{row}"] = "X" if prof.get('estudios') == "Sí" else ""
            sheet[f"U{row}"] = "X" if prof.get('estudios') == "No" else ""

            # Turno (Mañana, Tarde, Sab-Dom en V, W, X)
            if prof.get('turno') == "{Mañana}":
                sheet[f"V{row}"] = "X"
            elif prof.get('turno') == "{Tarde}":
                sheet[f"W{row}"] = "X"
            elif prof.get('turno') == "{Sab-Dom}":
                sheet[f"X{row}"] = "X"

            # Grado que atiende en la columna Y
            sheet[f"Y{row}"] = prof.get('nivel_grado')

            # Codificación en Z-AA-AE
            if prof.get('codificacion') in codificaciones:
                sheet[f"{codificaciones[prof['codificacion']]}{row}"] = "X"

            # Categoría en AF-AK
            if prof.get('categoria') in categorias:
                sheet[f"{categorias[prof['categoria']]}{row}"] = "X"

            # Obtener asistencias, inasistencias y justificaciones
            num_asistencias = self.valor_o_cero(prof.get('numero_asistencias'))
            num_inasistencias = self.valor_o_cero(prof.get('numero_inasistencias'))
            num_justificaciones = self.valor_o_cero(prof.get('numero_justificaciones'))

            # Cálculo de porcentajes
            porcentaje_asistencia = round((num_asistencias / dias_rango) * 100, 2) if dias_rango > 0 else 0
            porcentaje_inasistencia = round((num_inasistencias / dias_rango) * 100, 2) if dias_rango > 0 else 0

            porcentaje_justificadas = round((num_justificaciones / num_inasistencias) * 100, 2) if num_inasistencias > 0 else 0
            porcentaje_injustificadas = round(100 - porcentaje_justificadas, 2) if num_inasistencias > 0 else 0

            # Guardar valores en el archivo
            sheet[f"AL{row}"] = num_asistencias  # Número de asistencias
            sheet[f"AM{row}"] = f"{porcentaje_asistencia}%"  # Porcentaje de asistencias
            sheet[f"AN{row}"] = num_inasistencias  # Número de inasistencias
            sheet[f"AO{row}"] = f"{porcentaje_inasistencia}%"  # Porcentaje de inasistencias
            sheet[f"AP{row}"] = f"{porcentaje_justificadas}%"  # Porcentaje de justificaciones
            sheet[f"AQ{row}"] = f"{porcentaje_injustificadas}%"  # Porcentaje de injustificadas

            # Número de justificaciones en la columna AU
            sheet[f"AU{row}"] = num_justificaciones

            # Obtener el número de niñas y niños atendidos
            ninas = self.valor_o_cero(prof.get("ninas_atendidas"))
            ninos = self.valor_o_cero(prof.get("ninos_atendidos"))
            total_ninos = ninas + ninos  # Ahora esto no dará error

            # Guardar valores en el archivo
            sheet[f"AR{row}"] = ninas  # Cantidad de niñas
            sheet[f"AS{row}"] = ninos  # Cantidad de niños
            sheet[f"AT{row}"] = total_ninos  # Total de niños atendidos

            # Colocar "UEN Ciudad Miranda" en la columna D si hay texto en la columna B
            if sheet[f"B{row}"].value:
                sheet[f"D{row}"] = "UEN Ciudad Miranda"

        # Guardar el archivo modificado
        wb.save(self.output_path)

        # Llamar a la función para rellenar la segunda hoja
        self.rellenar_plantilla_excel_segundahoja(fecha_fin)

    def rellenar_plantilla_excel_segundahoja(self, fecha_fin):
        # Obtener datos de profesores y estudiantes
        # Convertir fecha_fin a tipo date
        fecha_fin_date = fecha_fin.date() if isinstance(fecha_fin, datetime) else fecha_fin

        # Obtener datos de estudiantes
        estudiantes_data = self.obtener_estudiantes_matriculados(fecha_fin_date)
        print("DEBUG: Datos obtenidos de la base de datos:", estudiantes_data)

        estudiantes_asistencias = self.obtener_estudiantes_asistencias(fecha_fin_date)

        # Convertir la lista de asistencias en un diccionario
        asistencias_dict = {}
        for id_asistencia, nombre_grado_asis, genero_asis, cantidad_asis in estudiantes_asistencias:
            if nombre_grado_asis not in asistencias_dict:
                asistencias_dict[nombre_grado_asis] = {}
            asistencias_dict[nombre_grado_asis][genero_asis] = cantidad_asis

        # Abrir el archivo de la plantilla
        wb = openpyxl.load_workbook(self.output_path)
        sheet = wb["MOVIMIENTO ESTADISTICO"]  # Usar la hoja "MOVIMIENTO ESTADISTICO"


                # Inicializar contadores
        inicial_varones_activos = 0
        inicial_hembras_activas = 0
        inicial_varones_ingresados = 0
        inicial_hembras_ingresadas = 0
        inicial_varones_egresados = 0
        inicial_hembras_egresadas = 0

        primaria_varones_activos = 0
        primaria_hembras_activas = 0
        primaria_varones_ingresados = 0
        primaria_hembras_ingresadas = 0
        primaria_varones_egresados = 0
        primaria_hembras_egresadas = 0

        # Inicializar contadores para Grado Inicial
        inicial_varones_activos = 0
        inicial_hembras_activas = 0
        inicial_varones_ingresados = 0
        inicial_hembras_ingresadas = 0

        # Inicializar contadores para Primer Grado
        primer_varones_activos = 0
        primer_hembras_activas = 0
        primer_varones_ingresados = 0
        primer_hembras_ingresadas = 0

        # Inicializar contadores para Segundo Grado
        segundo_varones_activos = 0
        segundo_hembras_activas = 0
        segundo_varones_ingresados = 0
        segundo_hembras_ingresadas = 0

        # Inicializar contadores para Tercer Grado
        tercer_varones_activos = 0
        tercer_hembras_activas = 0
        tercer_varones_ingresados = 0
        tercer_hembras_ingresadas = 0

        # Inicializar contadores para Cuarto Grado
        cuarto_varones_activos = 0
        cuarto_hembras_activas = 0
        cuarto_varones_ingresados = 0
        cuarto_hembras_ingresadas = 0

        # Inicializar contadores para Quinto Grado
        quinto_varones_activos = 0
        quinto_hembras_activas = 0
        quinto_varones_ingresados = 0
        quinto_hembras_ingresadas = 0

        # Inicializar contadores para Sexto Grado
        sexto_varones_activos = 0
        sexto_hembras_activas = 0
        sexto_varones_ingresados = 0
        sexto_hembras_ingresadas = 0

        # Procesar datos de estudiantes
        for estudiante in estudiantes_data:
            nombre_grado = estudiante[0]
            genero = estudiante[1]
            estado = estudiante[2]
            cantidad = estudiante[3]
            
            # Solo contar estudiantes con estado "Activo" o "Ingresado"
            if estado == "Activo":
                # Grado Inicial
                if nombre_grado == "Grado Inicial":
                    if genero == "varon":
                        inicial_varones_activos += cantidad
                    elif genero == "hembra":
                        inicial_hembras_activas += cantidad

                # Educación Primaria
                elif nombre_grado == "Primer Grado":
                    if genero == "varon":
                        primer_varones_activos += cantidad
                    elif genero == "hembra":
                        primer_hembras_activas += cantidad

                elif nombre_grado == "Segundo Grado":
                    if genero == "varon":
                        segundo_varones_activos += cantidad
                    elif genero == "hembra":
                        segundo_hembras_activas += cantidad

                elif nombre_grado == "Tercer Grado":
                    if genero == "varon":
                        tercer_varones_activos += cantidad
                    elif genero == "hembra":
                        tercer_hembras_activas += cantidad

                elif nombre_grado == "Cuarto Grado":
                    if genero == "varon":
                        cuarto_varones_activos += cantidad
                    elif genero == "hembra":
                        cuarto_hembras_activas += cantidad

                elif nombre_grado == "Quinto Grado":
                    if genero == "varon":
                        quinto_varones_activos += cantidad
                    elif genero == "hembra":
                        quinto_hembras_activas += cantidad

                elif nombre_grado == "Sexto Grado":
                    if genero == "varon":
                        sexto_varones_activos += cantidad
                    elif genero == "hembra":
                        sexto_hembras_activas += cantidad



            elif estado == "Ingresado":
                # Grado Inicial
                if nombre_grado == "Grado Inicial":
                    if genero == "varon":
                        inicial_varones_ingresados += cantidad
                    elif genero == "hembra":
                        inicial_hembras_ingresadas += cantidad

                # Educación Primaria
                elif nombre_grado == "Primer Grado":
                    if genero == "varon":
                        primer_varones_ingresados += cantidad
                    elif genero == "hembra":
                        primer_hembras_ingresadas += cantidad

                elif nombre_grado == "Segundo Grado":
                    if genero == "varon":
                        segundo_varones_ingresados += cantidad
                    elif genero == "hembra":
                        segundo_hembras_ingresadas += cantidad

                elif nombre_grado == "Tercer Grado":
                    if genero == "varon":
                        tercer_varones_ingresados += cantidad
                    elif genero == "hembra":
                        tercer_hembras_ingresadas += cantidad

                elif nombre_grado == "Cuarto Grado":
                    if genero == "varon":
                        cuarto_varones_ingresados += cantidad
                    elif genero == "hembra":
                        cuarto_hembras_ingresadas += cantidad

                elif nombre_grado == "Quinto Grado":
                    if genero == "varon":
                        quinto_varones_ingresados += cantidad
                    elif genero == "hembra":
                        quinto_hembras_ingresadas += cantidad

                elif nombre_grado == "Sexto Grado":
                    if genero == "varon":
                        sexto_varones_ingresados += cantidad
                    elif genero == "hembra":
                        sexto_hembras_ingresadas += cantidad
                    # Grado Inicial
            elif nombre_grado == "Grado Inicial":
                if genero == "varon" and estado == "Activo":
                    inicial_varones_activos += cantidad
                elif genero == "hembra" and estado == "Activo":
                    inicial_hembras_activas += cantidad
                elif genero == "varon" and estado == "Ingresado":
                    inicial_varones_ingresados += cantidad
                elif genero == "hembra" and estado == "Ingresado":
                    inicial_hembras_ingresadas += cantidad
                elif genero == "varon" and estado == "Egresado":
                    inicial_varones_egresados += cantidad
                elif genero == "hembra" and estado == "Egresado":
                    inicial_hembras_egresadas += cantidad

            # Educación Primaria
            elif nombre_grado in ["Primer Grado", "Segundo Grado", "Tercer Grado", "Cuarto Grado", "Quinto Grado", "Sexto Grado"]:
                if genero == "varon" and estado == "Activo":
                    primaria_varones_activos += cantidad
                elif genero == "hembra" and estado == "Activo":
                    primaria_hembras_activas += cantidad
                elif genero == "varon" and estado == "Ingresado":
                    primaria_varones_ingresados += cantidad
                elif genero == "hembra" and estado == "Ingresado":
                    primaria_hembras_ingresadas += cantidad
                elif genero == "varon" and estado == "Egresado":
                    primaria_varones_egresados += cantidad
                elif genero == "hembra" and estado == "Egresado":
                    primaria_hembras_egresadas += cantidad


        # Colocar los datos en las posiciones específicas
            # Colocar los datos en las posiciones específicas
        # Grado Inicial
        sheet["G14"] = inicial_varones_activos
        sheet["H14"] = inicial_hembras_activas
        sheet["G15"] = inicial_varones_ingresados
        sheet["H15"] = inicial_hembras_ingresadas
        sheet["G17"] = inicial_varones_egresados
        sheet["H17"] = inicial_hembras_egresadas

        # Educación Primaria
        sheet["K14"] = primaria_varones_activos
        sheet["L14"] = primaria_hembras_activas
        sheet["K15"] = primaria_varones_ingresados
        sheet["L15"] = primaria_hembras_ingresadas
        sheet["K17"] = primaria_varones_egresados
        sheet["L17"] = primaria_hembras_egresadas
        # Grado Inicial
        sheet["C25"] = inicial_varones_activos  # Total varones activos
        sheet["D25"] = inicial_hembras_activas  # Total hembras activas
        sheet["C26"] = asistencias_dict.get("Grado Inicial", {}).get("varon", 0)  # Asistencias varones
        sheet["D26"] = asistencias_dict.get("Grado Inicial", {}).get("hembra", 0)  # Asistencias hembras

        # Primer Grado
        sheet["E25"] = primer_varones_activos  # Total varones Primer Grado
        sheet["F25"] = primer_hembras_activas  # Total hembras Primer Grado
        sheet["E26"] = asistencias_dict.get("Primer Grado", {}).get("varon", 0) if "Primer Grado" in asistencias_dict else 0  # Asistencias varones
        sheet["F26"] = asistencias_dict.get("Primer Grado", {}).get("hembra", 0) if "Primer Grado" in asistencias_dict else 0  # Asistencias hembras

        # Segundo Grado
        sheet["G25"] = segundo_varones_activos  # Total varones Segundo Grado
        sheet["H25"] = segundo_hembras_activas  # Total hembras Segundo Grado
        sheet["G26"] = asistencias_dict.get("Segundo Grado", {}).get("varon", 0) if "Segundo Grado" in asistencias_dict else 0  # Asistencias varones
        sheet["H26"] = asistencias_dict.get("Segundo Grado", {}).get("hembra", 0) if "Segundo Grado" in asistencias_dict else 0  # Asistencias hembras

        # Tercer Grado
        sheet["I25"] = tercer_varones_activos  # Total varones Tercer Grado
        sheet["J25"] = tercer_hembras_activas  # Total hembras Tercer Grado
        sheet["I26"] = asistencias_dict.get("Tercer Grado", {}).get("varon", 0) if "Tercer Grado" in asistencias_dict else 0  # Asistencias varones
        sheet["J26"] = asistencias_dict.get("Tercer Grado", {}).get("hembra", 0) if "Tercer Grado" in asistencias_dict else 0  # Asistencias hembras

        # Cuarto Grado
        sheet["K25"] = cuarto_varones_activos  # Total varones Cuarto Grado
        sheet["L25"] = cuarto_hembras_activas  # Total hembras Cuarto Grado
        sheet["K26"] = asistencias_dict.get("Cuarto Grado", {}).get("varon", 0) if "Cuarto Grado" in asistencias_dict else 0  # Asistencias varones
        sheet["L26"] = asistencias_dict.get("Cuarto Grado", {}).get("hembra", 0) if "Cuarto Grado" in asistencias_dict else 0  # Asistencias hembras

        # Quinto Grado
        sheet["M25"] = quinto_varones_activos  # Total varones Quinto Grado
        sheet["N25"] = quinto_hembras_activas  # Total hembras Quinto Grado
        sheet["M26"] = asistencias_dict.get("Quinto Grado", {}).get("varon", 0) if "Quinto Grado" in asistencias_dict else 0  # Asistencias varones
        sheet["N26"] = asistencias_dict.get("Quinto Grado", {}).get("hembra", 0) if "Quinto Grado" in asistencias_dict else 0  # Asistencias hembras

        # Sexto Grado
        sheet["O25"] = sexto_varones_activos  # Total varones Sexto Grado
        sheet["P25"] = sexto_hembras_activas  # Total hembras Sexto Grado
        sheet["O26"] = asistencias_dict.get("Sexto Grado", {}).get("varon", 0) if "Sexto Grado" in asistencias_dict else 0  # Asistencias varones
        sheet["P26"] = asistencias_dict.get("Sexto Grado", {}).get("hembra", 0) if "Sexto Grado" in asistencias_dict else 0  # Asistencias hembras

            # Obtener datos de profesores
    # Obtener datos de profesores
        profesores_data = self.obtener_profesores(fecha_fin_date)

        # Inicializar contadores para profesores
        directores = 0
        sub_directores = 0
        coordinadores = 0
        docentes_inicial = 0
        docentes_primaria = 0
        auxiliares = 0
        docentes_fisica = 0
        profesores_por_horas = 0
        otros = 0
        personal_administrativo = 0
        obreros = 0
        vigilantes = 0
        musica = 0

        # Procesar datos de profesores
        for profesor in profesores_data:
            nombre_grado = profesor[0]  # Suponiendo que el nombre del grado está en la primera posición
            cargo = profesor[1]  # Suponiendo que el cargo está en la segunda posición
            cantidad = profesor[2]  # Suponiendo que la cantidad está en la tercera posición

            if cargo == "Director":
                directores += cantidad
            elif cargo == "Sub-Director":
                sub_directores += cantidad
            elif cargo == "Coordinador":
                coordinadores += cantidad
            elif nombre_grado == "Grado Inicial":
                docentes_inicial += cantidad
            elif nombre_grado in ["Primer Grado", "Segundo Grado", "Tercer Grado", "Cuarto Grado", "Quinto Grado", "Sexto Grado"]:
                docentes_primaria += cantidad
            elif cargo == "Auxiliar":
                auxiliares += cantidad
            elif cargo == "Docente Educación Física":
                docentes_fisica += cantidad
            elif cargo == "Profesor por Horas":
                profesores_por_horas += cantidad
            elif cargo == "Otros":
                otros += cantidad
            elif cargo == "Personal Administrativo":
                personal_administrativo += cantidad
            elif cargo == "Obrero":
                obreros += cantidad
            elif cargo == "Vigilante":
                vigilantes += cantidad
            elif cargo == "Música":
                musica += cantidad

        # Colocar los datos de profesores en las posiciones específicas
        sheet["B46"] = directores
        sheet["C46"] = sub_directores
        sheet["D46"] = coordinadores
        sheet["E46"] = docentes_inicial
        sheet["F46"] = docentes_primaria
        sheet["G46"] = auxiliares
        sheet["H46"] = docentes_fisica
        sheet["I46"] = profesores_por_horas
        sheet["J46"] = otros
        sheet["K46"] = personal_administrativo
        sheet["L46"] = obreros
        sheet["M46"] = vigilantes
        sheet["N46"] = musica

        # Guardar el archivo modificado
        wb.save(self.output_path)

    def obtener_datos_reporte(self, fecha_inicio, fecha_fin):
        conexion = self.conectar()
        
        if not conexion:
            print("❌ No se pudo establecer conexión con la base de datos.")
            return []

        try:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM obtener_datos_reporte(%s, %s);", (fecha_inicio, fecha_fin))
                columnas = [desc[0] for desc in cursor.description]
                resultados = []

                for fila in cursor.fetchall():
                    fila_dict = dict(zip(columnas, fila))

                    # Convertir "Cargo" en una lista si viene como string separada por comas
                    if "Cargo" in fila_dict and isinstance(fila_dict["Cargo"], str):
                        fila_dict["Cargo"] = fila_dict["Cargo"].split(",")

                    resultados.append(fila_dict)

                
                return resultados

        except Exception as e:
            print(f"⚠️ Error al obtener datos del reporte: {e}")
            return []

        finally:
            self.cerrar_conexion(conexion)

    def obtener_profesores(self, fecha_hasta):
        conexion = self.conectar()
        if not conexion:
            print("❌ No se pudo establecer conexión con la base de datos.")
            return []

        try:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM obtener_profesores(%s);", (fecha_hasta,))
                return cursor.fetchall()

        except Exception as e:
            print(f"⚠️ Error al obtener datos de profesores: {e}")
            return []

        finally:
            self.cerrar_conexion(conexion)

    def obtener_estudiantes_matriculados(self, fecha_hasta):
        conexion = self.conectar()
       
        if not conexion:
            print("❌ No se pudo establecer conexión con la base de datos.")
            return []

        try:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM obtener_estudiantes_matriculados(%s);", (fecha_hasta,))
                
                return cursor.fetchall()

        except Exception as e:
            print(f"⚠️ Error al obtener datos de estudiantes matriculados: {e}")
            return []

        finally:
            self.cerrar_conexion(conexion)

    def obtener_estudiantes_asistencias(self, fecha_hasta):
        conexion = self.conectar()
       
        if not conexion:
            print("❌ No se pudo establecer conexión con la base de datos.")
            return []

        try:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT * FROM obtener_asistencias_estudiantes(%s);", (fecha_hasta,))
                
                return cursor.fetchall()

        except Exception as e:
            print(f"⚠️ Error al obtener datos de estudiantes matriculados: {e}")
            return []

        finally:
            self.cerrar_conexion(conexion)

    @staticmethod
    def conectar():
        try:
            connection = psycopg2.connect(
                host="localhost",
                database="BASE_EUNB",  # Cambia esto al nombre de tu base de datos
                user="zek3rdc",  # Cambia a tu usuario de PostgreSQL
                password="prueba12P$A"  # Cambia a la contraseña de tu PostgreSQL
            )
            if connection:
                return connection
        except OperationalError as e:
            print(f"OperationalError al conectar a PostgreSQL: {e}")
            return None

    @staticmethod
    def cerrar_conexion(connection):
        if connection is not None:
            try:
                # Intentamos hacer una consulta simple para verificar si la conexión está activa
                connection.cursor().execute('SELECT 1')
                connection.close()  # Cierra la conexión a PostgreSQL
                print("Conexión cerrada exitosamente.")
            except Exception as e:
                print(f"Error al cerrar la conexión: {e}")

    def generar_reporte_excel(self, fecha_inicio, fecha_fin):
        """
        Obtiene los datos de la base de datos y genera el archivo Excel con los datos.

        :param fecha_inicio: Fecha de inicio del rango (formato 'YYYY-MM-DD').
        :param fecha_fin: Fecha de fin del rango (formato 'YYYY-MM-DD').
        """
        print(f"📥 Obteniendo datos del {fecha_inicio} al {fecha_fin}...")
        data = self.obtener_datos_reporte(fecha_inicio, fecha_fin)

        if not data:
            print("⚠️ No se encontraron datos en el rango de fechas especificado.")
            return

        print("📊 Datos obtenidos correctamente. Generando Excel...")

        # Convertir fechas a formato datetime para pasarlas a la función de Excel
        fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")

        self.rellenar_plantilla_excel(data, fecha_inicio=fecha_inicio_dt, fecha_fin=fecha_fin_dt)

        print(f"✅ Reporte generado exitosamente: {self.output_path}")

